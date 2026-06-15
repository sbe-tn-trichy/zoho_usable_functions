import os
import csv
import logging
import openpyxl
from datetime import datetime
from typing import Any, Dict, List, Optional
from ..core.config import Config
from ..core.exceptions import LedgerParsingError
from ..core.models import DotDict

logger = logging.getLogger(__name__)

def amt(v: Any) -> float:
    """Helper to convert amount values into floats cleanly."""
    try:
        return round(float(str(v or 0).replace(",", "")), 2)
    except (ValueError, TypeError):
        return 0.0

def parse_date(s: str) -> str:
    """Helper to parse GSTR-2B dates DD/MM/YYYY into ISO YYYY-MM-DD."""
    try:
        return datetime.strptime(s.strip(), "%d/%m/%Y").date().isoformat()
    except (ValueError, TypeError):
        return s.strip()

def load_gstr2b_csv(file_path: str) -> List[Dict[str, Any]]:
    """
    Loads and parses a GSTR-2B CSV file.
    
    Args:
        file_path (str): Path to the local consolidated GSTR-2B CSV file.
        
    Returns:
        List[Dict[str, Any]]: Normalized GSTR-2B transaction records.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"GSTR-2B CSV file not found: {file_path}")

    gst_rows = []
    with open(file_path, newline="", encoding="utf-8-sig") as fh:
        for r in csv.DictReader(fh):
            gst_rows.append({
                "gstin":         r["GSTIN of supplier"].strip(),
                "supplier":      r["Trade/Legal name"].strip(),
                "doc_type":      r["Document Type"].strip(),
                "doc_number":    r["Document Number"].strip(),
                "doc_date":      parse_date(r["Document Date"]),
                "doc_value":     amt(r["Document Value (₹)"]),
                "taxable_value": amt(r["Taxable Value (₹)"]),
                "igst":          amt(r["Integrated Tax(₹)"]),
                "cgst":          amt(r["Central Tax(₹)"]),
                "sgst":          amt(r["State/UT Tax(₹)"]),
                "itc":           r.get("ITC Availability", "").strip(),
            })
    return gst_rows

def parse_gstr2_report(file_path: str) -> Dict[str, Dict[str, Any]]:
    """
    Parses a downloaded Zoho Books GSTR-2 Inward Supplies report (.xlsx).
    
    Args:
        file_path (str): Path to the downloaded Excel file.
        
    Returns:
        Dict[str, Dict[str, Any]]: Document numbers mapped to details.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"GSTR-2 report file not found: {file_path}")

    wb = openpyxl.load_workbook(file_path)
    books_docs = {}
    
    # Process B2B (Invoices) and DN (Credit/Debit Notes)
    for sheet_name in ["b2b", "dn"]:
        if sheet_name not in wb.sheetnames:
            continue
        sheet = wb[sheet_name]
        for r in range(3, sheet.max_row + 1):
            doc_num = sheet.cell(r, 3).value
            if not doc_num:
                continue
            doc_num = str(doc_num).strip()
            
            vendor_name = str(sheet.cell(r, 2).value or "").strip()
            invoice_val = amt(sheet.cell(r, 5).value)
            taxable_val = amt(sheet.cell(r, 8).value)
            
            cgst = amt(sheet.cell(r, 11).value)
            sgst = amt(sheet.cell(r, 10).value)
            igst = amt(sheet.cell(r, 12).value)
            tax_val = cgst + sgst + igst
            
            if doc_num not in books_docs:
                books_docs[doc_num] = {
                    "vendor_name": vendor_name,
                    "sub_total": 0.0,
                    "tax_total": 0.0,
                    "total": invoice_val
                }
            books_docs[doc_num]["sub_total"] += taxable_val
            books_docs[doc_num]["tax_total"] += tax_val
            
    return books_docs

def reconcile_gstr2b_with_books(
    books_client: Optional[Any] = None,
    gstr2b_csv_path: str = "input_files/gst/gstr2b_reconciliation_consolidated.csv",
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    amount_tolerance: float = 1.0,
    temp_xlsx_path: Optional[str] = None
) -> Dict[str, Any]:
    """
    Performs full reconciliation between GSTR-2B CSV and Zoho Books inward supplies.
    Handles contact mapping resolution, report downloading, Excel parsing, and matching.
    
    Args:
        books_client (Any, optional): ZohoBooksAPI client instance. Auto-initialized if None.
        gstr2b_csv_path (str): Path to local GSTR-2B CSV.
        from_date (str, optional): From date in format YYYY-MM-DD. Computed from ledger if omitted.
        to_date (str, optional): To date in format YYYY-MM-DD. Computed from ledger if omitted.
        amount_tolerance (float): Difference allowance in ₹.
        temp_xlsx_path (str, optional): Filepath to save downloaded Zoho report.
        
    Returns:
        Dict[str, Any]: Result buckets (matched, discrepant, missing).
    """
    if not books_client:
        from ..core.auth import get_books_client
        books_client = get_books_client()

    logger.info(f"Loading GSTR-2B CSV from {gstr2b_csv_path}...")
    if os.path.isdir(gstr2b_csv_path):
        import glob
        csv_files = sorted(glob.glob(os.path.join(gstr2b_csv_path, "*.csv")))
        if not csv_files:
            raise LedgerParsingError(f"No GSTR-2B CSV files found in directory: {gstr2b_csv_path}")
        gst_rows = []
        for file in csv_files:
            logger.info(f"Loading CSV: {file}")
            gst_rows.extend(load_gstr2b_csv(file))
    else:
        gst_rows = load_gstr2b_csv(gstr2b_csv_path)

    # Date computation in src/
    if not from_date or not to_date:
        valid_dates = [g["doc_date"] for g in gst_rows if g.get("doc_date")]
        if not valid_dates:
            raise LedgerParsingError("No valid document dates found in GSTR-2B file to compute date range.")
        if not from_date:
            min_date_str = min(valid_dates)
            min_dt = datetime.strptime(min_date_str, "%Y-%m-%d")
            from_date = min_dt.replace(day=1).date().isoformat()
        if not to_date:
            import calendar
            max_date_str = max(valid_dates)
            max_dt = datetime.strptime(max_date_str, "%Y-%m-%d")
            last_day = calendar.monthrange(max_dt.year, max_dt.month)[1]
            to_date = max_dt.replace(day=last_day).date().isoformat()
        logger.info(f"Computed date range from GSTR-2B ledger: {from_date} to {to_date}")

    if not temp_xlsx_path:
        dt = datetime.strptime(from_date, "%Y-%m-%d")
        month_name = dt.strftime("%B").lower()
        temp_xlsx_path = f"input_files/gst/inward_supplies_{month_name}_{dt.year}.xlsx"
    
    invoices = [g for g in gst_rows if g["doc_type"] != "Credit Note"]
    cred_notes = [g for g in gst_rows if g["doc_type"] == "Credit Note"]
    
    # Resolve unknown GSTINs dynamically
    gstin_to_vendor_id = dict(Config.GSTIN_TO_VENDOR_ID)
    for g in gst_rows:
        gstin = g["gstin"]
        if gstin and (gstin not in gstin_to_vendor_id or gstin_to_vendor_id[gstin] is None):
            logger.info(f"Resolving unknown GSTIN {gstin} via Zoho Books...")
            res = books_client.request('GET', 'contacts', params={'search_text': gstin})
            contacts = res.get('contacts', [])
            contact_id = contacts[0]['contact_id'] if contacts else None
            gstin_to_vendor_id[gstin] = contact_id
            if contact_id:
                logger.info(f"Resolved {gstin} -> {contacts[0]['contact_name']}")

    # Download GSTR-2 report
    logger.info(f"Downloading GSTR-2 Inward Supplies report from Zoho Books to {temp_xlsx_path}...")
    dirname = os.path.dirname(temp_xlsx_path)
    if dirname:
        os.makedirs(dirname, exist_ok=True)
    books_client.gst.download_gstr_inward_supplies(
        save_path=temp_xlsx_path,
        params={
            "from_date": from_date,
            "to_date": to_date,
            "filter_by": "TransactionDate.CustomDate",
            "tax_settings_id": Config.ZOHO_TAX_SETTINGS_ID,
            "response_option": "1",
            "x-zb-source": "zbclient",
            "accept": "xlsx",
            "file_name": "Summary of Inward Supplies (GSTR-2)"
        }
    )
    
    # Fallback to output/ directory path if saved there by SDK
    actual_xlsx_path = temp_xlsx_path
    if not os.path.isabs(actual_xlsx_path):
        out_path = os.path.abspath(os.path.join("output", temp_xlsx_path))
        if os.path.exists(out_path):
            actual_xlsx_path = out_path

    logger.info(f"Parsing GSTR-2 report: {actual_xlsx_path}...")
    books_docs = parse_gstr2_report(actual_xlsx_path)

    matched_inv = []
    discrepant_inv = []
    missing_inv = []
    matched_cn = []
    discrepant_cn = []
    missing_cn = []

    for g in gst_rows:
        vendor_id = gstin_to_vendor_id.get(g["gstin"])
        is_cn = g["doc_type"] == "Credit Note"

        if not vendor_id:
            (missing_cn if is_cn else missing_inv).append({"gst": g, "reason": "Vendor not mapped in Books"})
            continue

        rec = books_docs.get(g["doc_number"])
        if not rec:
            (missing_cn if is_cn else missing_inv).append({"gst": g, "reason": "Not found in Zoho Books"})
            continue

        b_sub = amt(rec.get("sub_total"))
        b_tax = amt(rec.get("tax_total"))
        b_total = amt(rec.get("total"))
        
        if b_sub == 0 and b_total > 0:
            b_sub = b_total - b_tax

        gst_tax = round(g["igst"] + g["cgst"] + g["sgst"], 2)
        tv_diff = round(b_sub - g["taxable_value"], 2)
        tax_diff = round(b_tax - gst_tax, 2)

        entry = {
            "gst":        g,
            "books":      rec,
            "b_sub":      b_sub,
            "b_tax":      b_tax,
            "b_total":    b_total,
            "gst_tax":    gst_tax,
            "tv_diff":    tv_diff,
            "tax_diff":   tax_diff,
        }

        bucket_ok = matched_cn if is_cn else matched_inv
        bucket_bad = discrepant_cn if is_cn else discrepant_inv

        if abs(tv_diff) <= amount_tolerance:
            bucket_ok.append(entry)
        else:
            bucket_bad.append(entry)

    return DotDict({
        "matched_invoices": matched_inv,
        "discrepant_invoices": discrepant_inv,
        "missing_invoices": missing_inv,
        "matched_credits": matched_cn,
        "discrepant_credits": discrepant_cn,
        "missing_credits": missing_cn,
        "gst_rows_count": len(gst_rows),
        "invoices_count": len(invoices),
        "credits_count": len(cred_notes),
        "gstin_to_vendor_id": gstin_to_vendor_id
    })


def clean_gstr2b_xlsx(xlsx_path: str, csv_path: str) -> None:
    """
    Parses a GSTR-2B Excel file (.xlsx) and writes its contents to a GSTR-2B CSV format.
    Looks for sheets named 'B2B' and 'CDNR' (case-insensitive).
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    
    headers = [
        "GSTIN of supplier",
        "Trade/Legal name",
        "Document Type",
        "Document Number",
        "Document Date",
        "Document Value (₹)",
        "Taxable Value (₹)",
        "Integrated Tax(₹)",
        "Central Tax(₹)",
        "State/UT Tax(₹)",
        "ITC Availability"
    ]
    
    rows = []
    
    # Process sheets B2B and CDNR
    for sheet_name in wb.sheetnames:
        name_lower = sheet_name.lower()
        if "b2b" not in name_lower and "cdnr" not in name_lower:
            continue
            
        sheet = wb[sheet_name]
        
        # Find header row
        header_row = -1
        header_map = {}
        
        # Scan first 15 rows for the headers
        for r in range(1, 16):
            row_vals = [str(sheet.cell(r, c).value or "").strip().lower() for c in range(1, 30)]
            if any("gstin" in val for val in row_vals):
                header_row = r
                # Map column indices (1-based) to headers
                for c in range(1, 30):
                    val = str(sheet.cell(r, c).value or "").strip().lower()
                    if not val:
                        continue
                    
                    if "gstin" in val:
                        header_map["gstin"] = c
                    elif "trade/legal" in val or ("legal" in val and "name" in val):
                        header_map["supplier"] = c
                    elif "invoice number" in val or "note number" in val or "document number" in val or "invoice no" in val or "note no" in val:
                        header_map["doc_number"] = c
                    elif "invoice date" in val or "note date" in val or "document date" in val or "invoice dt" in val or "note dt" in val:
                        header_map["doc_date"] = c
                    elif "invoice value" in val or "note value" in val or "document value" in val:
                        header_map["doc_value"] = c
                    elif "taxable value" in val:
                        header_map["taxable_value"] = c
                    elif "integrated tax" in val:
                        header_map["igst"] = c
                    elif "central tax" in val:
                        header_map["cgst"] = c
                    elif "state/ut tax" in val or "state tax" in val:
                        header_map["sgst"] = c
                    elif "itc availability" in val or "itc elig" in val:
                        header_map["itc"] = c
                    elif "document type" in val or "note type" in val:
                        header_map["doc_type"] = c
                break
                
        if header_row == -1:
            logger.warning(f"Could not find GSTR-2B header row in sheet: {sheet_name}")
            continue
            
        # Read data rows
        for r in range(header_row + 1, sheet.max_row + 1):
            gstin = str(sheet.cell(r, header_map.get("gstin", 1)).value or "").strip()
            if not gstin or len(gstin) < 3 or gstin.lower() == "total":
                continue
                
            # Document Type default based on sheet name
            if "doc_type" in header_map:
                doc_type_val = str(sheet.cell(r, header_map["doc_type"]).value or "").strip()
            else:
                doc_type_val = "Credit Note" if "cdnr" in name_lower else "Invoice"
                
            # Date formatting (if date object or string)
            raw_date = sheet.cell(r, header_map.get("doc_date", 4)).value
            if isinstance(raw_date, datetime):
                doc_date_str = raw_date.strftime("%d/%m/%Y")
            elif raw_date:
                # If date is YYYY-MM-DD string, parse and format as DD/MM/YYYY
                try:
                    dt = datetime.strptime(str(raw_date).strip(), "%Y-%m-%d")
                    doc_date_str = dt.strftime("%d/%m/%Y")
                except ValueError:
                    doc_date_str = str(raw_date).strip()
            else:
                doc_date_str = ""
                
            supplier = str(sheet.cell(r, header_map.get("supplier", 2)).value or "").strip()
            doc_number = str(sheet.cell(r, header_map.get("doc_number", 3)).value or "").strip()
            doc_value = str(sheet.cell(r, header_map.get("doc_value", 5)).value or "").strip()
            taxable_value = str(sheet.cell(r, header_map.get("taxable_value", 6)).value or "").strip()
            igst = str(sheet.cell(r, header_map.get("igst", 7)).value or "0").strip()
            cgst = str(sheet.cell(r, header_map.get("cgst", 8)).value or "0").strip()
            sgst = str(sheet.cell(r, header_map.get("sgst", 9)).value or "0").strip()
            itc = str(sheet.cell(r, header_map.get("itc", 10)).value or "Y").strip()
            
            rows.append({
                "GSTIN of supplier": gstin,
                "Trade/Legal name": supplier,
                "Document Type": doc_type_val,
                "Document Number": doc_number,
                "Document Date": doc_date_str,
                "Document Value (₹)": doc_value,
                "Taxable Value (₹)": taxable_value,
                "Integrated Tax(₹)": igst,
                "Central Tax(₹)": cgst,
                "State/UT Tax(₹)": sgst,
                "ITC Availability": itc
            })
            
    # Write to CSV
    os.makedirs(os.path.dirname(csv_path), exist_ok=True)
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)

